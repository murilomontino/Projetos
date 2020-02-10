from openpyxl import load_workbook
from openpyxl.styles import Font


arquivo_excel  = load_workbook("planilha/SOLIM2016DIAGNOSTICO.xlsx")
planilha_leitura = arquivo_excel.active


linha : int = 4483
total : int = 14719
paciente : list = []
coluna_usavel : int = 9

estilo = Font(name='Arial', size='8')

while linha <= total:
    codigo = planilha_leitura.cell(row=linha, column=1).value
    if isinstance( codigo, (int, float) ):
        paciente += [linha]
        coluna_usavel = 9
    else:
        diagnostico = planilha_leitura.cell(row=linha, column=8).value
        _cell = planilha_leitura.cell(row=paciente[-1], column=coluna_usavel, value=diagnostico)
        _cell.font = estilo
        planilha_leitura.cell(row=linha, column=8, value='')
        coluna_usavel += 1
    linha += 1
else:
    print('Terminou')
arquivo_excel.save("planilha_completa/SOLIM2016DIAGNOSTICO.xlsx")

exit()




